#coding=utf-8
import sys
import os
import openpyxl
base_path = "data.xlsx" #项目存放读取文件的路径，注意不要和我的弄错了
# base_path = os.getcwd()
# print(base_path)
sys.path.append(base_path)
class HandleExcel:
    def load_excel(self):
        '''
        加载excel
        '''
        open_excel = openpyxl.load_workbook(base_path)#拿到excel的所有内容
        return open_excel
    def get_sheet_data(self,index=None):
        '''
        加载所有sheet的内容
        '''
        sheet_name = self.load_excel().sheetnames#拿到sheetnames的所有内容
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data
    def get_cell_value(self,row,cols):
        '''
        获取某一个单元格内容
        '''
        data = self.get_sheet_data().cell(row=row,column=cols)
        return data
    def get_rows(self):
        row = self.get_sheet_data().max_row
        return row
    def get_rows_value(self,row):
        '''
        获取某一行的内容
        '''
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def getProjectCode(self,str):
        strs = str.split("】")
        if len(strs) < 3:
            return ""
        return strs[2].strip("【")

    def save(self,dict, dictUser):
        workbook = self.load_excel()

        if "统计结果" in workbook.sheetnames:
            sheet = workbook["统计结果"]
        else:
            sheet = workbook.create_sheet("统计结果")
        sheet.delete_rows(1,sheet.max_row)
        sheet.append(["项目编号", "总工时", "总投入人数"])

        for projectCode in dict:
            sheet.append({'A': projectCode, 'B': dict[projectCode],'C': len(dictUser[projectCode])})
        workbook.save(base_path)
if __name__ == '__main__':
    handle = HandleExcel()

    dict = {}
    dictUser = {}
    for i in  range(2,handle.get_rows()):
        str = handle.get_rows_value(i)[8]
        projectCode = handle.getProjectCode(str)
        if projectCode == "":
            continue
        workTime = float(handle.get_rows_value(i)[7])
        user = handle.get_rows_value(i)[0]
        if projectCode in dict:
            dict[projectCode] += workTime
            dictUser[projectCode].add(user)
            continue
        dict[projectCode] = workTime
        dictUser[projectCode] = {user}

    handle.save(dict=dict, dictUser=dictUser)
